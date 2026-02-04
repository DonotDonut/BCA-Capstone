from methods.database import Database

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import psycopg2



class Report:
  @staticmethod
  def create_asia_report_table(db_parameters):
      conn = Database.get_connection(db_parameters)

      try:
          with conn.cursor() as cur:
              # Get columns for airline_routes
              cur.execute("""
                  SELECT column_name
                  FROM information_schema.columns
                  WHERE table_schema = 'public'
                    AND table_name = 'airline_routes';
              """)
              routes_cols = {r[0] for r in cur.fetchall()}

              # Pick the best available join from airline_routes -> aircraft
              join_sql = None

              if "aircraft_id" in routes_cols:
                  join_sql = "ON ac.aircraft_id = r.aircraft_id"
              elif "iata_code" in routes_cols:
                  join_sql = "ON ac.iata_code = r.iata_code"
              elif "icao_code" in routes_cols:
                  join_sql = "ON ac.icao_code = r.icao_code"
              elif "equipment" in routes_cols:
                  join_sql = "ON (ac.iata_code = r.equipment OR ac.icao_code = r.equipment)"

              aircraft_join_clause = (
                  f"LEFT JOIN aircraft ac {join_sql}"
                  if join_sql
                  else "LEFT JOIN aircraft ac ON 1=0"
              )

              cur.execute("DROP TABLE IF EXISTS asia_report;")

              create_sql = f"""
              CREATE TABLE asia_report AS
              SELECT
                r.airline_id,
                r.airline_code,
                COALESCE(al.name, '(unknown)') AS airline_name,

                -- Asia -> non-Asia
                COUNT(*) FILTER (
                  WHERE r.source_in_asia = TRUE AND r.dest_in_asia = FALSE
                ) AS flghts_out_of_asia,

                -- non-Asia -> Asia
                COUNT(*) FILTER (
                  WHERE r.source_in_asia = FALSE AND r.dest_in_asia = TRUE
                ) AS flghts_in_asia,

                -- Asia -> Asia
                COUNT(*) FILTER (
                  WHERE r.source_in_asia = TRUE AND r.dest_in_asia = TRUE
                ) AS flghts_within_asia,

                -- total includes out + in + within (touches Asia)
                COUNT(*) FILTER (
                  WHERE r.source_in_asia = TRUE OR r.dest_in_asia = TRUE
                ) AS total_flights_to_asia,

                -- Passenger capacity sums
                SUM(COALESCE(ac.seat_capacity, 0)) FILTER (
                  WHERE r.source_in_asia = TRUE AND r.dest_in_asia = FALSE
                ) AS pax_out_of_asia,

                SUM(COALESCE(ac.seat_capacity, 0)) FILTER (
                  WHERE r.source_in_asia = FALSE AND r.dest_in_asia = TRUE
                ) AS pax_in_asia,

                -- Asia -> Asia passengers
                SUM(COALESCE(ac.seat_capacity, 0)) FILTER (
                  WHERE r.source_in_asia = TRUE AND r.dest_in_asia = TRUE
                ) AS pax_within_asia,

                -- pax total includes out + in + within (touches Asia)
                SUM(COALESCE(ac.seat_capacity, 0)) FILTER (
                  WHERE r.source_in_asia = TRUE OR r.dest_in_asia = TRUE
                ) AS pax_total_to_asia

              FROM airline_routes r
              LEFT JOIN airlines al
                ON al.airline_id = r.airline_id
              {aircraft_join_clause}

              GROUP BY r.airline_id, r.airline_code, al.name

              -- Keep airlines that have ANY Asia-related flight (in/out/within)
              HAVING COUNT(*) FILTER (
                WHERE r.source_in_asia = TRUE OR r.dest_in_asia = TRUE
              ) > 0;
              """

              cur.execute(create_sql)

          conn.commit()
          print("asia_report table created successfully.")
          if not join_sql:
              print("Warning: No aircraft link column found in airline_routes; pax_* columns will be 0.")
          else:
              print(f"Aircraft join used: {join_sql}")
              
          # Load table into DataFrame
          df = pd.read_sql("SELECT * FROM asia_report;", conn)

          # Export to Excel
          df.to_excel("output data/asia_report.xlsx", index=False)

      except Exception as e:
          conn.rollback()
          print("Error creating asia_report:", e)

      finally:
          conn.close()
            
    
  @staticmethod
  def hex_to_argb(hex_color: str) -> str:
      h = hex_color.lstrip("#").upper()
      return "FF" + h

  @staticmethod
  def apply_airline_highlights(excel_path: str):
      airline_colors = {
          "China Southern Airlines": "#1f77b4",
          "China Eastern Airlines":  "#ff7f0e",
          "Air China":               "#2ca02c",
          "Shenzhen Airlines":       "#d62728",
          "Turkish Airlines":        "#9467bd",
          "All Nippon Airways":      "#8c564b",
          "Hainan Airlines":         "#e377c2",
          "Sichuan Airlines":        "#7f7f7f",
          "Air India Limited":       "#bcbd22",
          "Xiamen Airlines":         "#17becf",
      }

      wb = load_workbook(excel_path)
      ws = wb.active

      header_row = 1
      headers = {}
      for col in range(1, ws.max_column + 1):
          val = ws.cell(row=header_row, column=col).value
          if isinstance(val, str) and val.strip():
              headers[val.strip()] = col

      if "airline_name" not in headers:
          wb.close()
          raise KeyError("Column 'airline_name' not found in the exported Excel file.")

      airline_name_col = headers["airline_name"]

      for r in range(header_row + 1, ws.max_row + 1):
          name = ws.cell(row=r, column=airline_name_col).value
          if isinstance(name, str) and name in airline_colors:
              argb = Report.hex_to_argb(airline_colors[name])
              fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
              ws.cell(row=r, column=airline_name_col).fill = fill

      wb.save(excel_path)
      wb.close()



  def get_airlines_using_top10_airports(db_parameters):
      conn = Database.get_connection(db_parameters)
      sql = """
      WITH top_airports AS (
          SELECT airport_id, name, iata, total_in_out
          FROM airports
          ORDER BY total_in_out DESC
          LIMIT 10
      ),
      routes_touching_top AS (
          SELECT r.airline_id, r.source_airport_id AS airport_id
          FROM airline_routes r
          JOIN top_airports ta ON ta.airport_id = r.source_airport_id

          UNION ALL

          SELECT r.airline_id, r.dest_airport_id AS airport_id
          FROM airline_routes r
          JOIN top_airports ta ON ta.airport_id = r.dest_airport_id
      ),
      airline_usage AS (
          SELECT
              airport_id,
              airline_id,
              COUNT(*) AS route_records_touching_airport
          FROM routes_touching_top
          WHERE airline_id IS NOT NULL
          GROUP BY airport_id, airline_id
      )
      SELECT
          ta.airport_id,
          ta.iata AS airport_iata,
          ta.name AS airport_name,
          ta.total_in_out,
          al.airline_id,
          al.name AS airline_name,
          al.iata AS airline_iata,
          al.icao AS airline_icao,
          au.route_records_touching_airport
      FROM airline_usage au
      JOIN top_airports ta ON ta.airport_id = au.airport_id
      LEFT JOIN airlines al ON al.airline_id = au.airline_id
      ORDER BY
          ta.total_in_out DESC,
          ta.airport_id,
          au.route_records_touching_airport DESC,
          airline_name;
      """

      excel_path = "output data/top_airports_in_asia_report.xlsx"

      try:
          df = pd.read_sql(sql, conn)
          df.to_excel(excel_path, index=False)

          Report.apply_airline_highlights(excel_path)

      finally:
          conn.close()
          
  def get_airlines_unique_airport_counts(db_parameters):
        conn = Database.get_connection(db_parameters)

        sql = """
        WITH airline_airports AS (

            -- Airports an airline DEPARTS from
            SELECT
                airline_id,
                source_airport_id AS airport_id
            FROM airline_routes
            WHERE airline_id IS NOT NULL

            UNION

            -- Airports an airline ARRIVES at
            SELECT
                airline_id,
                dest_airport_id AS airport_id
            FROM airline_routes
            WHERE airline_id IS NOT NULL
        ),

        airline_unique_counts AS (
            SELECT
                airline_id,
                COUNT(DISTINCT airport_id) AS unique_airports_touched
            FROM airline_airports
            GROUP BY airline_id
        )

        SELECT
            al.airline_id,
            al.name AS airline_name,
            al.iata AS airline_iata,
            al.icao AS airline_icao,
            auc.unique_airports_touched
        FROM airline_unique_counts auc
        LEFT JOIN airlines al ON al.airline_id = auc.airline_id
        ORDER BY unique_airports_touched DESC, airline_name;
        """

        excel_path = "output data/airlines_unique_airports_report.xlsx"

        try:
            df = pd.read_sql(sql, conn)
            df.to_excel(excel_path, index=False)

            Report.apply_airline_highlights(excel_path)

        finally:
            conn.close()
